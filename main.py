import pandas as pd
import openpyxl
import json
from datetime import datetime


class bank:
    
    def __init__(self):
        print("초기 설정 완료")
        self.log = dict()
        self.log['server_log'] = list()
        self.data = dict()
        self.price = {1 : 1000, 2 : 800, 3 : 600, 4 : 550, 5 : 500, 6 : 400, 7 : 350, 8 : 300, 9 : 100, 10 : 50}
        self.bank_data = pd.Series([self.data.keys(),self.data.values()], index = ['name', 'inform'])
        
        try:    
            self.files = openpyxl.load_workbook('data.xlsx')
            self.xlsx = self.files.active
            self.load()

        except FileNotFoundError:
            files = openpyxl.Workbook()
            files.save('data.xlsx')            
            
            self.files = openpyxl.load_workbook('data.xlsx')
            self.xlsx = self.files.active
            
    def add_user(self, user):
        try:
            if self.data[user][0] != -1:
                print("이미 존제하는 유저입니다.")
                return 1

            elif self.data[user][0] == -1:
                print(f"{user} 추가 성공")
                self.data[user][0] = 0
                return 0

        except KeyError:
            print(f"{user} 추가 성공")
            self.data[user] = [0, 0, 'date', 0]
            return 0

    def manage(self, user, money):
        try:
            if self.data[user][0] != -1:
                print(f"{user} 한테서 {money} 추가했습니다.")
                self.data[user][0] += int(money)
                return 0
            
            else:
                print('계좌가 없는 유저입니다.')
                return 1
                
        except KeyError:
            print('계좌가 없는 유저입니다.')
            return 1

    def save(self):
        for i in range(len(self.data)): 
            self.xlsx.cell(row=i+2, column=1).value = list(self.data.keys())[i]
            self.xlsx.cell(row=i+2, column=2).value = self.data[list(self.data.keys())[i]][0]
            self.xlsx.cell(row=i+2, column=3).value = self.data[list(self.data.keys())[i]][1]
            self.xlsx.cell(row=i+2, column=4).value = self.data[list(self.data.keys())[i]][2]
            self.xlsx.cell(row=i+2, column=5).value = self.data[list(self.data.keys())[i]][3]

        self.files.save('data.xlsx')

    def load(self):
        num = 1
        while True:
            if self.xlsx.cell(row=num+1, column=1).value != None:
                self.data[self.xlsx.cell(row=num+1, column=1).value] = [self.xlsx.cell(row=num+1, column=2).value, self.xlsx.cell(row=num+1, column=3).value,self.xlsx.cell(row=num+1, column=4).value, self.xlsx.cell(row=num+1, column=5).value]
                num += 1

            else:
                break
            
        self.bank_data = pd.Series([self.data.keys(),self.data.values()], index = ['name', 'inform'])

        try: 
            with open('log.json', 'r', encoding='UTF-8-sig') as f:
                self.log = json.load(f)

        except FileNotFoundError:
            pass

    
    def buy(self, user_id, item):
        if self.data[user_id][0] < self.price[item]:
            return 1  

        elif self.data[user_id][0] >= self.price[item]:
            if item == 1:
                self.data[user_id][1] += 1
            self.data[user_id][0] -= self.price[item]
            return 0

    def attend(self, user_id):
        try:
            if self.data[user_id][2] == str(datetime.today().strftime("%Y/%m/%d")):
                print('이미 출석 했습니다')
                return 1

            elif self.data[user_id][2] != str(datetime.today().strftime("%Y/%m/%d")):
                self.data[user_id][2] = str(datetime.today().strftime("%Y/%m/%d"))
                # print(self.data[user_id])
                # print(type(self.data[user_id][3]))
                self.data[user_id][3] += 1
                print(str(datetime.today().strftime("%Y/%m/%d")))
                if self.data[user_id][0] != -1:
                    self.data[user_id][0] += 5
                    print('5코인 추가.')
                    return 2
                     
                return 0

        except KeyError:
            print('없는 유저 입니다.')
            self.data[user_id] = [-1, 0, str(datetime.today().strftime("%Y/%m/%d")), 1]
            return 0


    def cheack(self, user):
        return self.data[user][0]

    def log_server_add(self, log_data):
        self.log['server_log'].append(log_data + '\n')

    def log_channel_add(self, ch, msg):
        try:
            self.log[ch].append(msg)
            
        except KeyError:
            self.log[ch] = list()
            self.log[ch].append(msg)

    def log_save(self):
        with open('log.json', 'w', encoding='UTF-8-sig') as f:
            f.write(json.dumps(self.log, ensure_ascii=False, indent=4))
        
            
    
        


# bot = bank()
# bot.add_user('책스쵸코')
# bot.add_user('잭스')
# print(bot.bank_data)
# bot.manage('이누', 100)
# print(bot.bank_data)
# bot.attend(10)





