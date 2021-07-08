import openpyxl
import json
from datetime import datetime

class bank:

    def __init__(self):
        self.log = dict()
        self.log['server'] = list()
        self.data = dict()
        self.goods = {1 : 1000, 2 : 800, 3 : 600, 4 : 550, 5 : 500, 6 : 400, 7 : 350, 8 : 300, 9 : 100, 10 : 50}

        self.files = openpyxl.load_workbook('data.xlsx')
        self.xlsx = self.files.active
        self.load()

    def new_bank_user(self, user):
        try:
            if self.data[user][0] == -1:
                self.data[user][0] = 0
                return 0

            elif self.data[user][0] == -1:
                return 1

        except KeyError:
            self.new_user(user)
            self.new_bank_user(user)
            
    def new_user(self, user):
        self.data[user] = [-1, 0, str(datetime.today().strftime("%Y-%m-%d")), 0, str(datetime.today().strftime("%Y-%m-%d")), 0]

    def manage(self, user, money):
        try:
            self.data[user][0] = self.data[user][0] + money
            return 0

        except KeyError:
            return 1

    def buy(self, user, i_num):
        try:
            if self.data[user][0] >= self.goods[i_num]:
                self.data[user][0] = self.data[user][0] - self.goods[i_num]
                if i_num == 1:
                    self.data[user][1] += 1
                return 0
            
            else:
                return 0

        except KeyError:
            return -1

    def cheack(self, user):
        return self.data[user][0]

    def attend(self, user_id):
        # try:
        if self.data[user_id][2] == str(datetime.today().strftime("%Y/%m/%d")):
            print('이미 출석 했습니다')
            return 2

        elif self.data[user_id][2] != str(datetime.today().strftime("%Y/%m/%d")):
            self.data[user_id][2] = str(datetime.today().strftime("%Y/%m/%d"))
            self.data[user_id][3] += 1
            print(str(datetime.today().strftime("%Y/%m/%d")))
            if self.data[user_id][0] != -1:
                self.data[user_id][0] += 10
                print('10코인 추가.')
                return 1
                    
            return 0

        # except KeyError:
        #     print('없는 유저 입니다.')
        #     self.data[user_id] = [-1, 0, 1,str(datetime.today().strftime("%Y/%m/%d")), str(datetime.today().strftime("%Y/%m/%d")),0]
        #     return 0

    def save_data(self):
        for i in range(0, len(self.data)): 
            self.xlsx.cell(row=i+2, column=1).value = list(self.data.keys())[i]
            for l in range(0, 6):
                self.xlsx.cell(row=i+2, column=l+2).value = self.data[list(self.data.keys())[i]][l]
                
        self.files.save('data.xlsx')

    def load(self):
        num = 1
        while True:
            if self.xlsx.cell(row=num+1, column=1).value != None:
                user = self.xlsx.cell(row=num+1, column=1).value
                coin = int(self.xlsx.cell(row=num+1, column=2).value)
                warning_down = int(self.xlsx.cell(row=num+1, column=3).value)
                att_date = self.xlsx.cell(row=num+1, column=4).value
                att_count = int(self.xlsx.cell(row=num+1, column=5).value)
                first_ac = self.xlsx.cell(row=num+1, column=6).value
                chat_count = int(self.xlsx.cell(row=num+1, column=7).value)
                self.data[user] = [coin, warning_down, att_date, att_count, first_ac, chat_count]
                num += 1

            else:
                break
        
        try: 
            with open('log.json', 'r', encoding='UTF-8-sig') as f:
                self.log = json.load(f)

        except FileNotFoundError:
            pass
    
    def log_server_add(self, log_data):
        self.log['server_log'].append(log_data + '\n')

    def log_channel_user_add(self, ch, msg, user):
        try:
            self.log[ch].append(msg)

        except KeyError:
            self.log[ch] = list()
            self.log[ch].append(msg)

        try:
            self.data[user][5] = self.data[user][5] + 1

        except KeyError:
            self.new_user(user) 
    
    def log_save(self):
        with open('log.json', 'w', encoding='UTF-8-sig') as f:
            f.write(json.dumps(self.log, ensure_ascii=False, indent=4))
        
    def conversion(self):
        for filename in self.log.keys():
            txt = ''

            for log in self.log[filename]:
                txt += log

            with open(f'{filename}.txt', 'w', encoding='UTF-8-sig') as f:
                f.write(txt)


# bot = bank()
